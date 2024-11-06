from os import path, remove
from openpyxl import load_workbook
from src.config_system.config_service import ConfigService
from src.config_system.utility import Utility


class TransferService:

    def __init__(self) -> None:
        self.config_path = Utility.get_config_path()
        self.path_to_temp_declaration = Utility.get_temp_path()

    def transfer(self) -> None:
        config_data = ConfigService.static_read_config(self.config_path)
        registry_article_key = config_data["article_data"]["registry_article_key"]
        declaration_article_key = config_data["article_data"]["declaration_article_key"]

        registry_wb = load_workbook(
            path.normpath(config_data["paths"]["registry_path"]))
        declaration_wb = load_workbook(
            path.normpath(config_data["paths"]["declaration_path"]))

        registry_sheet = registry_wb.active
        declaration_sheet = declaration_wb.active

        registry_articles = dict()
        for registry_row in range(int(config_data["start_rows"]["registry_start_row"]), registry_sheet.max_row + 1):
            registry_article_cell = str(
                registry_sheet[registry_article_key + str(registry_row)].value).rstrip(" ")
            if registry_article_cell:
                registry_articles[registry_article_cell] = registry_row

        for declaration_row in range(int(config_data["start_rows"]["declaration_start_row"]), declaration_sheet.max_row + 1):
            declaration_article_cell = str(
                declaration_sheet[declaration_article_key + str(declaration_row)].value).rstrip(" ")
            if declaration_article_cell in registry_articles.keys():
                registry_row = registry_articles[declaration_article_cell]
                for registry_sign in config_data["transfer_pairs"].keys():
                    declaration_sheet[config_data["transfer_pairs"][registry_sign] + str(
                        declaration_row)] = registry_sheet[registry_sign + str(registry_row)].value

        registry_wb.close()
        declaration_wb.save(path.normpath(self.path_to_temp_declaration))
        declaration_wb.close()

    def open_temp_declaration(self) -> None:
        Utility.start_file(self.path_to_temp_declaration)

    def save_temp_declaration(self, save_path: str) -> None:
        file = load_workbook(
            path.normpath(self.path_to_temp_declaration))
        file.save(path.normpath(save_path))
        file.close()

    def delete_temp_declaration(self) -> None:
        remove(path.normpath(self.path_to_temp_declaration))
