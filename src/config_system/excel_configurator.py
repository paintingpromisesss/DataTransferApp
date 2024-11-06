from openpyxl import Workbook, load_workbook
from os import remove
from tempfile import NamedTemporaryFile
from re import match
from src.config_system.excel_styles import data_cell_style, other_cell_style, article_cell_style
from openpyxl.styles import Protection
from src.config_system.utility import Utility
from src.config_system.config_service import ConfigService


class ExcelConfigurator:
    def __init__(self) -> None:
        self.config_path = Utility.get_config_path()
    def init_temp_file(self):
        self.temp_file = NamedTemporaryFile(
            suffix=".xlsx", delete=False
        )
        self.config_data = ConfigService.static_read_config(self.config_path)

        book = Workbook()
        book.save(self.temp_file.name)
        book.close()
        wb = load_workbook(self.temp_file.name)

        wb.add_named_style(data_cell_style)
        wb.add_named_style(other_cell_style)
        wb.add_named_style(article_cell_style)

        self._apply_styles_to_cells(wb)
        self._fill_excel_with_config_data(wb)
        wb.save(self.temp_file.name)
        wb.close()
        self.temp_file.close()


    def _apply_styles_to_cells(self, wb: Workbook) -> None:
        sheet = wb.active
        sheet.protection.sheet = True
        try:
            article_cell = self.config_data["article_data"]["registry_article_key"]
            sheet[article_cell + "1"].style = "article_cell_style"
        except Exception:
            pass

        for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=100):
            for cell in col:
                if cell.style != "article_cell_style":
                    cell.style = "data_cell_style"
                    cell.protection = Protection(locked=False)

        for col in sheet.iter_cols(min_row=2, max_row=100, min_col=1, max_col=100):
            for cell in col:
                cell.style = "other_cell_style"

        for col in sheet.iter_cols(min_row=1, max_row=100, min_col=101, max_col=200):
            for cell in col:
                cell.style = "other_cell_style"

    def _fill_excel_with_config_data(self, wb: Workbook) -> None:
        sheet = wb.active
        for key in self.config_data["transfer_pairs"]:
            sheet[key + "1"] = self.config_data["transfer_pairs"][key]

        try:
            sheet[self.config_data["article_data"]["registry_article_key"] +
                  "1"] = self.config_data["article_data"]["declaration_article_key"]
        except Exception:
            pass

    def open_excel_configurator(self) -> None:
        self.init_temp_file()
        Utility.start_file(self.temp_file.name)
        if Utility.is_file_closed(self.temp_file.name):
            wb = load_workbook(self.temp_file.name)
            sheet = wb.active
            for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=100):
                key = col[0].column_letter
                value = col[0].value
                if (key != self.config_data["article_data"]["registry_article_key"]) and (
                        value != self.config_data["article_data"]["declaration_article_key"]):
                    if value is not None:
                        if bool(match(r'[A-Za-z]', value)) and len(value) == 1:
                            self.config_data["transfer_pairs"][key] = value.upper(
                            )
                    else:
                        try:
                            self.config_data["transfer_pairs"].pop(key)
                        except KeyError:
                            pass
            wb.close()
            ConfigService.static_edit_transfer_pairs(
                config_path=self.config_path, edited_config=self.config_data
            )
        self.delete_temp_file()

    def delete_temp_file(self) -> None:
        remove(self.temp_file.name)
