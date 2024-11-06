from openpyxl.styles import Side, NamedStyle, Border, Alignment, PatternFill, Font

bd = Side(style="thin", color="000000")
data_cell_style = NamedStyle(name="data_cell_style", border=Border(
    left=bd, right=bd, top=bd, bottom=bd), alignment=Alignment(horizontal="center"), fill=PatternFill(fill_type="solid", fgColor="DDEBF7"), font=Font(name="Calibri", size=11, bold=True))

other_cell_style = NamedStyle(name="other_cell_style", fill=PatternFill(
    fill_type="solid", fgColor="E5EDEF"))

article_cell_style = NamedStyle(name="article_cell_style", fill=PatternFill(
    fill_type="solid", fgColor="C00000"), font=Font(name="Calibri", size=11, bold=True))
